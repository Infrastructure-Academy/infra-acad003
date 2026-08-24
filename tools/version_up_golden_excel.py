from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import json
import shutil

root = Path('/home/ubuntu/infra-acad003')
source = root / 'docs/resources/files/GE-ISI_Master_v1_14Jul26.xlsx'
out = root / 'docs/resources/files/GE-ISI_Master_v2_25Aug26.xlsx'
catalog_path = root / 'archive/2026-08-25/drizzle-schema-export/tables.json'
report_path = root / 'archive/2026-08-25/50-table-schema-report.md'

catalog = json.loads(catalog_path.read_text())
tables = catalog['tables']
shutil.copy2(source, out)
wb = load_workbook(out)
if 'Schema 50 Reconciliation' in wb.sheetnames:
    del wb['Schema 50 Reconciliation']
ws = wb.create_sheet('Schema 50 Reconciliation', 0)
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'B8'
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 11
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 29
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 13
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 44

navy = '1F4E79'
blue = 'D6E3F0'
gold = 'C9A227'
light = 'F7F9FB'
red = 'FCE4D6'
green = 'E2F0D9'
thin = Side(style='thin', color='D9E2F3')

ws.merge_cells('B2:H2')
ws['B2'] = 'Drizzle Schema — 50-Table Reconciliation'
ws['B2'].font = Font(name='Georgia', size=18, bold=True, color=navy)
ws['B2'].alignment = Alignment(horizontal='left')
ws['B3'] = 'Version'
ws['C3'] = 'v2 — 25 Aug 2026'
ws['E3'] = 'Source workbook'
ws['F3'] = 'GE-ISI_Master_v1_14Jul26.xlsx'
ws['B4'] = 'Verified declarations'
ws['C4'] = len(tables)
ws['E4'] = 'Reported total'
ws['F4'] = 50
ws['B5'] = 'Unresolved against restored schema'
ws['C5'] = 50 - len(tables)
ws['E5'] = 'Database rows included'
ws['F5'] = 'No'
for cell in ('B3','E3','B4','E4','B5','E5'):
    ws[cell].font = Font(name='Calibri', bold=True, color=navy)
for cell in ('C3','F3','C4','F4','C5','F5'):
    ws[cell].font = Font(name='Calibri', color='000000')
ws.merge_cells('B6:H6')
ws['B6'] = 'The Telegram manifest reports 50 tables. Direct inspection of the restored Memorial Drizzle source verified 19 mysqlTable declarations. The remaining 31 are listed as unresolved; no absent table names or data are fabricated.'
ws['B6'].alignment = Alignment(wrap_text=True, vertical='top')
ws['B6'].font = Font(name='Calibri', italic=True, color='666666')
ws.row_dimensions[6].height = 34

headers = ['Report slot', 'Status', 'Export symbol', 'SQL table', 'Source line', 'Columns', 'Evidence / action']
for col, header in enumerate(headers, start=2):
    c = ws.cell(row=8, column=col, value=header)
    c.font = Font(name='Georgia', bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=navy)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

rows = []
for idx, table in enumerate(tables, start=1):
    rows.append([idx, 'VERIFIED', table['exportName'], table['tableName'], table['sourceLine'], len(table['columns']), 'Definition present in restored drizzle/schema.ts'])
for idx in range(len(tables) + 1, 51):
    rows.append([idx, 'UNRESOLVED', '—', '—', '—', '—', 'No matching definition in restored schema; obtain the corresponding schema revision or authorized metadata export'])

for r, row in enumerate(rows, start=9):
    for c_idx, value in enumerate(row, start=2):
        cell = ws.cell(row=r, column=c_idx, value=value)
        cell.font = Font(name='Calibri', size=10, color='000000')
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = Border(bottom=thin)
        if r % 2 == 0:
            cell.fill = PatternFill('solid', fgColor=light)
    status = ws.cell(row=r, column=3)
    status.fill = PatternFill('solid', fgColor=green if row[1] == 'VERIFIED' else red)
    status.font = Font(name='Calibri', bold=True, color='006100' if row[1] == 'VERIFIED' else '9C0006')

ref = f'B8:H{8+len(rows)}'
tab = Table(displayName='Schema50Reconciliation', ref=ref)
tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)
ws.auto_filter.ref = ref
ws['B61'] = 'Notes'
ws['B61'].font = Font(name='Georgia', bold=True, color=navy)
ws.merge_cells('B62:H63')
ws['B62'] = 'This is a schema reconciliation report, not a live database backup. It contains table definitions and verification status only. The exact TypeScript source is archived beside the catalog in archive/2026-08-25/drizzle-schema-export/.'
ws['B62'].alignment = Alignment(wrap_text=True, vertical='top')
ws['B62'].font = Font(name='Calibri', italic=True, color='666666')
ws['B62'].comment = Comment('Generated from the restored Memorial project source; no credentials or database rows included.', 'Manus AI')

report_lines = [
    '# 50-Table Schema Report', '',
    '**Workbook version:** `GE-ISI_Master_v2_25Aug26.xlsx`  ',
    f'**Reported total:** 50  ',
    f'**Verified in restored Drizzle source:** {len(tables)}  ',
    f'**Unresolved:** {50 - len(tables)}  ',
    '',
    '## Verified table definitions', '',
    '| # | Export symbol | SQL table | Source line | Columns |',
    '|---:|---|---|---:|---:|',
]
for idx, table in enumerate(tables, start=1):
    report_lines.append(f"| {idx} | `{table['exportName']}` | `{table['tableName']}` | {table['sourceLine']} | {len(table['columns'])} |")
report_lines += [
    '',
    '## Reconciliation', '',
    'The Telegram manifest reports 50 database tables. Direct inspection of the restored Memorial project found 19 `mysqlTable` declarations in `drizzle/schema.ts` and no additional table constructors in `drizzle/relations.ts`. The workbook therefore records 31 unresolved slots without assigning invented names or data.',
    '',
    '## Scope limitation', '',
    'This report and workbook contain schema definitions only. They do not contain live database rows, credentials, or secrets. Resolving the remaining 31 requires the corresponding schema revision or an authorized database metadata export.',
    '',
    '## References', '',
    '[1]: https://github.com/Infrastructure-Academy/infra-acad003 "Infrastructure Academy repository"',
]
report_path.write_text('\n'.join(report_lines) + '\n')
wb.save(out)
print(f'Created {out} with {len(wb.sheetnames)} sheets and {len(tables)} verified tables')
print(f'Created {report_path}')
