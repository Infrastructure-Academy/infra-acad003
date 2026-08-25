from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/infra-acad003-live/docs/resources/files/GE-ISI_Master_v2_25Aug26.xlsx')
wb = load_workbook(path, read_only=False, data_only=False)
print('workbook', path.name)
print('sheet_count', len(wb.sheetnames))
print('new_tab_present', 'Schema 50 Reconciliation' in wb.sheetnames)
print('new_tab_index', wb.sheetnames.index('Schema 50 Reconciliation') + 1 if 'Schema 50 Reconciliation' in wb.sheetnames else None)
for name in ('INDEX', 'INDEX (Print)', 'Schema 50 Reconciliation'):
    if name not in wb.sheetnames:
        continue
    ws = wb[name]
    print(f'-- {name} --')
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is not None and any(term in str(value).lower() for term in ('human', 'holistic', 'haptic', 'icu', 'ice', 'iq', 'eq', 'cq', 'relay', 'dominion', 'machine', 'protocol', 'schema')):
                link = cell.hyperlink.location if cell.hyperlink and cell.hyperlink.location else (cell.hyperlink.target if cell.hyperlink else '')
                print(cell.coordinate, repr(str(value)[:300]), 'LINK=', repr(link))
